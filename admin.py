import streamlit as st
import pandas as pd
import os
import io
from datetime import datetime
from database import (get_all_users, get_all_books, get_all_requests, get_stats,
                       add_book, delete_book, update_request_status, get_books_read_count)
from library import GENRES, get_cover_html, genre_badge, stars_html
from database import get_avg_rating

def show_admin_dashboard():
    # Hard security check — never allow non-admins even if URL is manipulated
    user = st.session_state.get("user", {})
    if not user or user.get("role") != "admin":
        st.markdown('<div class="alert-error">🚫 Access denied. Admins only.</div>', unsafe_allow_html=True)
        st.stop()

    st.markdown("""
    <div style='display:flex; align-items:center; gap:0.75rem; margin-bottom:2rem;'>
        <div style='font-size:1.5rem;'>👑</div>
        <div style='font-family:Playfair Display,serif; font-size:1.8rem; font-weight:700; color:#c9a84c;'>Admin Dashboard</div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "  📊 Stats  ",
        "  📤 Upload Book  ",
        "  📚 Manage Books  ",
        "  👥 Users  ",
        "  📬 Requests  "
    ])

    with tab1:
        _show_stats()
    with tab2:
        _show_upload()
    with tab3:
        _show_manage_books()
    with tab4:
        _show_users()
    with tab5:
        _show_requests()


def _show_stats():
    stats = get_stats()
    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    cards = [
        (stats["total_books"], "Total Books", "📚"),
        (stats["total_users"], "Registered Users", "👥"),
        (stats["total_reviews"], "Reviews Written", "⭐"),
        (stats["pending_requests"], "Pending Requests", "📬"),
    ]
    for col, (num, label, icon) in zip([c1,c2,c3,c4], cards):
        with col:
            st.markdown(f"""
            <div class="stat-card">
                <div style="font-size:1.8rem;">{icon}</div>
                <div class="stat-number">{num}</div>
                <div class="stat-label">{label}</div>
            </div>
            """, unsafe_allow_html=True)

    if stats["top_books"]:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-header">🏆 Top Rated Books</div>', unsafe_allow_html=True)
        for i, book in enumerate(stats["top_books"], 1):
            avg = round(book["avg"] or 0, 1)
            stars_str = "★" * round(avg) + "☆" * (5 - round(avg))
            st.markdown(f"""
            <div class="review-card" style="display:flex; justify-content:space-between; align-items:center;">
                <div>
                    <span style="color:#c9a84c; font-weight:700; margin-right:0.75rem;">#{i}</span>
                    <span style="color:#f0e6d3;">{book['title']}</span>
                </div>
                <div>
                    <span style="color:#c9a84c;">{stars_str}</span>
                    <span style="color:#8b9bb4; font-size:0.85rem;"> {avg}/5 ({book['cnt']} reviews)</span>
                </div>
            </div>
            """, unsafe_allow_html=True)


def _show_upload():
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">📤 Upload New Book</div>', unsafe_allow_html=True)

    with st.form("upload_book_form"):
        c1, c2 = st.columns(2)
        with c1:
            title = st.text_input("Book Title *", placeholder="e.g. The Way of Kings")
            author = st.text_input("Author *", placeholder="e.g. Brandon Sanderson")
        with c2:
            genre = st.selectbox("Genre *", GENRES[1:])  # skip "All"
            
        description = st.text_area("Description", placeholder="A brief summary or blurb...", height=100)
        cover_file = st.file_uploader("Cover Image", type=["jpg","jpeg","png","webp"])
        book_file = st.file_uploader("Book File (PDF or EPUB) *", type=["pdf","epub"])

        st.markdown("<br>", unsafe_allow_html=True)
        submitted = st.form_submit_button("Upload Book →", use_container_width=True)

        if submitted:
            if not title.strip() or not author.strip():
                st.markdown('<div class="alert-error">⚠️ Title and Author are required.</div>', unsafe_allow_html=True)
            elif not book_file:
                st.markdown('<div class="alert-error">⚠️ Please upload a book file (PDF or EPUB).</div>', unsafe_allow_html=True)
            else:
                # Save files
                safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:50]
                ts = datetime.now().strftime("%Y%m%d%H%M%S")

                ext = book_file.name.split(".")[-1].lower()
                file_path = os.path.join("books", f"{safe_title}_{ts}.{ext}")
                with open(file_path, "wb") as f:
                    f.write(book_file.read())

                cover_path = None
                if cover_file:
                    cext = cover_file.name.split(".")[-1].lower()
                    cover_path = os.path.join("covers", f"{safe_title}_{ts}.{cext}")
                    with open(cover_path, "wb") as f:
                        f.write(cover_file.read())

                add_book(
                    title=title.strip(),
                    author=author.strip(),
                    genre=genre,
                    description=description.strip(),
                    cover_path=cover_path,
                    file_path=file_path,
                    uploaded_by=st.session_state.user["id"]
                )
                st.markdown(f'<div class="alert-success">✅ "{title}" uploaded successfully!</div>', unsafe_allow_html=True)
                st.rerun()


def _show_manage_books():
    st.markdown("<br>", unsafe_allow_html=True)
    books = get_all_books()

    if not books:
        st.markdown('<div class="alert-info">📭 No books in the library yet.</div>', unsafe_allow_html=True)
        return

    st.markdown(f"<div style='color:#8b9bb4; margin-bottom:1rem;'>{len(books)} books total</div>", unsafe_allow_html=True)

    for book in books:
        avg, cnt = get_avg_rating(book["id"])
        with st.expander(f"📚 {book['title']} — {book['author']}"):
            c1, c2 = st.columns([1, 3])
            with c1:
                cover_html = get_cover_html(book.get("cover_path"), height=160)
                st.markdown(cover_html, unsafe_allow_html=True)
            with c2:
                stars_str = "★" * round(avg) + "☆" * (5 - round(avg))
                st.markdown(f"""
                <div style='color:#f0e6d3;'><b>Genre:</b> {book['genre']}</div>
                <div style='color:#f0e6d3;'><b>Rating:</b> <span style='color:#c9a84c;'>{stars_str}</span> ({cnt} reviews)</div>
                <div style='color:#8b9bb4; font-size:0.85rem;'>Uploaded: {str(book.get('uploaded_at',''))[:10]}</div>
                <div style='color:#8b9bb4; font-size:0.85rem; margin-top:4px;'>{book.get('description','') or ''}</div>
                """, unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button(f"🗑 Delete '{book['title']}'", key=f"del_{book['id']}"):
                    delete_book(book["id"])
                    st.markdown(f'<div class="alert-success">🗑 Deleted successfully.</div>', unsafe_allow_html=True)
                    st.rerun()


def _show_users():
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">👥 All Users</div>', unsafe_allow_html=True)

    users = get_all_users()

    # Build DataFrame
    rows = []
    for u in users:
        rows.append({
            "Name": u["name"],
            "Email": u["email"],
            "Role": u["role"].upper(),
            "Joined Date": str(u.get("joined_at",""))[:10],
            "Last Login": str(u.get("last_login",""))[:10] if u.get("last_login") else "Never",
            "Books Read": get_books_read_count(u["id"])
        })

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Excel download
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="PageVault Users")
        ws = writer.sheets["PageVault Users"]
        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        gold_font = Font(color="C9A84C", bold=True, size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = gold_font
            cell.alignment = Alignment(horizontal="center")
        # Column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output.seek(0)
    st.download_button(
        label="⬇ Download Users as Excel",
        data=output.getvalue(),
        file_name=f"pagevault_users_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


def _show_requests():
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">📬 Book Requests</div>', unsafe_allow_html=True)

    requests = get_all_requests()

    if not requests:
        st.markdown('<div class="alert-info">📭 No requests yet.</div>', unsafe_allow_html=True)
        return

    pending = [r for r in requests if r["status"] == "pending"]
    fulfilled = [r for r in requests if r["status"] == "fulfilled"]

    st.markdown(f"<div style='color:#8b9bb4; margin-bottom:1rem;'>📌 {len(pending)} pending · ✅ {len(fulfilled)} fulfilled</div>", unsafe_allow_html=True)

    for req in requests:
        status_cls = "status-fulfilled" if req["status"] == "fulfilled" else "status-pending"
        with st.expander(f"{'✅' if req['status']=='fulfilled' else '📌'} {req['book_title']} — requested by {req['user_name']}"):
            st.markdown(f"""
            <div style='color:#f0e6d3; margin-bottom:0.5rem;'>
                <b>Book:</b> {req['book_title']} {f"by {req['author']}" if req.get('author') else ''}<br/>
                <b>Requested by:</b> {req['user_name']} ({req['user_email']})<br/>
                <b>Date:</b> {str(req.get('created_at',''))[:10]}<br/>
                <b>Status:</b> <span class="{status_cls}">{req['status'].upper()}</span>
            </div>
            {f'<div class="review-text"><b>Reason:</b> {req["reason"]}</div>' if req.get("reason") else ''}
            """, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                if req["status"] == "pending":
                    if st.button("✅ Mark Fulfilled", key=f"fulfill_{req['id']}", use_container_width=True):
                        update_request_status(req["id"], "fulfilled")
                        st.rerun()
            with c2:
                if req["status"] == "fulfilled":
                    if st.button("↩ Mark Pending", key=f"pend_{req['id']}", use_container_width=True):
                        update_request_status(req["id"], "pending")
                        st.rerun()